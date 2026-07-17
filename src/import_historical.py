#!/usr/bin/env python3
"""Jednorázový import historických dat ze starého Excel reportu do CSV."""
import sys
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
import metrics, data_store as store, hubspot_client as hs

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR  = REPO_ROOT / "data"

SHEET_TO_HUBSPOT_NAME = {
    "Martin":   "Martin Korbelar",
    "Lukáš":    "Lukas Hora",
    "Veronika": "Veronika Kincová",
    "Mykyta":   "Mykyta Artiukhov",
    "Richard":  "Richard Brůža",
    "Olda":     "Oldřich Huzil",
    "Julie":    "Julie Mrkvickova",
}
STAGE_LABELS = [
    "Lead Engaged","Meeting negotiation / contacted","Intro meeting agreed",
    "Awaiting confirmation / Not Now","Qualify","Discover","Validate",
    "Decide","Commit","Tech intro","Implementation","Testing",
]

def import_snapshots(wb):
    rows = []
    for sheet, hname in SHEET_TO_HUBSPOT_NAME.items():
        if sheet not in wb.sheetnames: print(f"  přeskakuji {sheet}"); continue
        ws = wb[sheet]
        last = 1
        for c in range(2, ws.max_column+1):
            v = ws.cell(16, c).value
            if isinstance(v, (int, float)): last = c
            else: break
        stage_rows = {}
        for r in range(4, 16):
            lbl = ws.cell(r,1).value
            if lbl: stage_rows[lbl.strip()] = r
        n = 0
        for c in range(2, last+1):
            wd = ws.cell(3,c).value
            if not isinstance(wd, datetime): continue
            wd = wd.date()
            # Starý excel má v hlavičce NEDĚLI – normalizuj na pondělí
            monday = wd - timedelta(days=wd.weekday())
            wl = metrics.week_label_for_date(monday); n += 1
            for stage in STAGE_LABELS:
                r = stage_rows.get(stage)
                amt = ws.cell(r,c).value if r else 0
                if not isinstance(amt,(int,float)): amt = 0
                rows.append({"week_label":wl,"week_monday":monday.isoformat(),
                    "deal_id":f"hist-{sheet}-{stage}-{c}".replace(" ","_"),
                    "deal_name":"","owner_id":"","owner_name":hname,
                    "stage_id":stage,"stage_label":stage,"amount":float(amt),
                    "closedate":monday.isoformat(),"pipeline_id":"",
                    "is_closed_won":False,"is_closed_lost":False})
        print(f"  {sheet}: {n} týdnů")
    return pd.DataFrame(rows, columns=store.SNAPSHOTS_COLUMNS)

def import_ledger(wb, token=None):
    cids = set(); raw = []
    for sheet, outcome in [("Won","won"),("Lost","lost")]:
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        for r in range(2, ws.max_row+1):
            cd = ws.cell(r,1).value; amt = ws.cell(r,4).value
            owner = ws.cell(r,5).value; did = ws.cell(r,7).value; cid = ws.cell(r,8).value
            if not did or not owner: continue
            # Přeskočit řádky před 1.1.2026
            if not isinstance(cd, datetime) or cd.date() < __import__("datetime").date(2026,1,1): continue
            cd_iso = cd.isoformat() if isinstance(cd,datetime) else ""
            cwl = metrics.week_label_for_date(cd.date()) if isinstance(cd,datetime) else ""
            raw.append({"deal_id":str(did),"deal_name":"","_cid":str(cid) if cid else "",
                "owner_name":owner,"closedate":cd_iso,"amount":float(amt) if isinstance(amt,(int,float)) else 0.0,
                "outcome":outcome,"week_label":cwl})
            if cid: cids.add(str(cid))
    cnames = {}
    if token and cids:
        print(f"  Dotahuji {len(cids)} názvů firem...")
        cnames = hs.get_company_names(token, list(cids))
    rows = []
    for r in raw:
        rows.append({"deal_id":r["deal_id"],"deal_name":r["deal_name"],
            "company_name":cnames.get(r["_cid"],""),
            "owner_name":r["owner_name"],"closedate":r["closedate"],
            "amount":r["amount"],"outcome":r["outcome"],"week_label":r["week_label"]})
    print(f"  Won: {sum(1 for r in rows if r['outcome']=='won')}, Lost: {sum(1 for r in rows if r['outcome']=='lost')}")
    return pd.DataFrame(rows, columns=store.LEDGER_COLUMNS)

def main():
    if len(sys.argv) < 2:
        print("Použití: python src/import_historical.py /cesta/k/Sales_reporting_2026.xlsx"); sys.exit(1)
    wb = load_workbook(sys.argv[1], data_only=True)
    token = None
    try: token = hs.load_token()
    except RuntimeError: print("  (bez HUBSPOT_TOKEN - company name zůstane prázdný)")
    print("Snapshoty..."); snap = import_snapshots(wb)
    print("Ledger...");    ledger = import_ledger(wb, token)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    sp = DATA_DIR/"deals_snapshots.csv"; lp = DATA_DIR/"deals_closed_ledger.csv"
    snap.to_csv(sp, index=False); ledger.to_csv(lp, index=False)
    print(f"\nHotovo:\n  {sp} ({len(snap)} řádků)\n  {lp} ({len(ledger)} řádků)")

if __name__ == "__main__":
    main()
