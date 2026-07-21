#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Jednorázová oprava poškozeného týdne W28 (2026-07-06—2026-07-12).

Ten týden se v živém provozu zapsal jako nekompletní (přerušený HubSpot
fetch) - většina fází spadla na 0. Tento skript ho nahradí správnými
hodnotami vytaženými ze dvou syrových exportů (stejný princip jako dřívější
jednorázová oprava W9-W27):

- HubSpot_Deals_By_Stage_2026.xlsx      = "do konce roku" (fixed cutoff)
- HubSpot_Deals_By_Stage_DYNAMIC_2026.xlsx = "Rolling 18" (18měsíční cutoff)

Nové fáze (Not a lead, To be contacted, Qualified -> Deal) a Won/Lost se
nepřenáší - stejně jako u W9-W27. Všechny ostatní týdny (W1-W27, W29+)
zůstávají nedotčené.

POUŽITÍ:
    Uprav proměnné FIXED_PATH / DYNAMIC_PATH / SNAPSHOTS_PATH níže na
    skutečné cesty ve svém repu, a spusť:
        python3 src/fix_w28_onetime.py
"""
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import data_store as store

# --- ZDE UPRAV CESTY, ať odpovídají tvému repu / staženým souborům ---
FIXED_PATH = "HubSpot_Deals_By_Stage_2026.xlsx"
DYNAMIC_PATH = "HubSpot_Deals_By_Stage_DYNAMIC_2026.xlsx"
SNAPSHOTS_PATH = Path("data/deals_snapshots.csv")
# ----------------------------------------------------------------------

CANONICAL_STAGES = [
    "Lead Engaged", "Meeting negotiation / contacted", "Intro meeting agreed",
    "Awaiting confirmation / Not Now", "Qualify", "Discover", "Validate",
    "Decide", "Commit", "Tech intro", "Implementation", "Testing",
]

SHEET_TO_TARGET_OWNER = {
    "Martin Korbelar": "Martin Korbelar",
    "Lukas Hora": "Lukas Hora",
    "Veronika Kincová": "Veronika Kincová",
    "Mykyta Artiukhov": "Mykyta Artiukhov",
    "Richard Brůža": "Richard Brůža",
    "Oldřich Huzil": "Oldřich Huzil",
    "Julie Mrkvičková": "Julie Mrkvickova",
    "Julie Mrkvickova": "Julie Mrkvickova",
    "Dominika Šudová": "Dominika Šudová",
}

TARGET_WEEK = 28


def extract_week_num(header):
    m = re.match(r"2026-W(\d+)", header or "")
    return int(m.group(1)) if m else None


def read_raw_file(path):
    """Vrací {(owner, stage, week_label): amount} jen pro TARGET_WEEK."""
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        target_owner = SHEET_TO_TARGET_OWNER.get(sheet)
        if not target_owner:
            continue
        ws = wb[sheet]
        stage_rows = {}
        for r in range(2, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if lbl:
                stage_rows[lbl.strip()] = r
        for c in range(2, ws.max_column + 1):
            header = ws.cell(1, c).value
            if extract_week_num(header) != TARGET_WEEK:
                continue
            for stage in CANONICAL_STAGES:
                r = stage_rows.get(stage)
                if r is None:
                    continue
                val = ws.cell(r, c).value
                amt = float(val) if isinstance(val, (int, float)) else 0.0
                key = (target_owner, stage, header)
                out[key] = out.get(key, 0.0) + amt
    return out


print("Načítám FIXED (do konce roku)...")
fixed = read_raw_file(FIXED_PATH)
print(f"  {len(fixed)} (owner, stage) záznamů pro W{TARGET_WEEK}")

print("Načítám DYNAMIC (Rolling 18)...")
dynamic = read_raw_file(DYNAMIC_PATH)
print(f"  {len(dynamic)} (owner, stage) záznamů pro W{TARGET_WEEK}")

all_keys = set(fixed.keys()) | set(dynamic.keys())
print(f"Celkem unikátních (owner, stage, week) pro W{TARGET_WEEK}: {len(all_keys)}")

new_rows = []
negative_deltas = 0
week_label_seen = None
for owner, stage, week_label in all_keys:
    week_label_seen = week_label
    fixed_amt = fixed.get((owner, stage, week_label), 0.0)
    dynamic_amt = dynamic.get((owner, stage, week_label), 0.0)
    delta = dynamic_amt - fixed_amt
    if delta < -0.01:
        negative_deltas += 1
        delta = 0.0

    monday = date.fromisocalendar(2026, TARGET_WEEK, 1)

    if fixed_amt != 0 or delta == 0:
        new_rows.append({
            "week_label": week_label, "week_monday": monday.isoformat(),
            "deal_id": f"histfix-{owner}-{stage}-{week_label}-A".replace(" ", "_"),
            "deal_name": "", "owner_id": "", "owner_name": owner,
            "stage_id": stage, "stage_label": stage, "amount": fixed_amt,
            "closedate": monday.isoformat(), "pipeline_id": "",
            "is_closed_won": False, "is_closed_lost": False,
        })
    if delta > 0:
        beyond_eoy_date = date(2027, 6, 1)
        new_rows.append({
            "week_label": week_label, "week_monday": monday.isoformat(),
            "deal_id": f"histfix-{owner}-{stage}-{week_label}-B".replace(" ", "_"),
            "deal_name": "", "owner_id": "", "owner_name": owner,
            "stage_id": stage, "stage_label": stage, "amount": delta,
            "closedate": beyond_eoy_date.isoformat(), "pipeline_id": "",
            "is_closed_won": False, "is_closed_lost": False,
        })

print(f"Záporné delty (dynamic < fixed, ořezáno na 0): {negative_deltas}")
print(f"Nově sestavených řádků pro W{TARGET_WEEK}: {len(new_rows)}")

if not new_rows:
    print("POZOR: nic se nenašlo pro W28 v uploadnutých souborech - zkontroluj "
          "názvy listů/cesty k souborům. Nic se nezapisuje.")
    sys.exit(1)

new_df = pd.DataFrame(new_rows, columns=store.SNAPSHOTS_COLUMNS)

existing = pd.read_csv(SNAPSHOTS_PATH)
existing_week_nums = existing["week_label"].apply(extract_week_num)
mask_to_remove = existing_week_nums == TARGET_WEEK
print(f"Odebírám {mask_to_remove.sum()} starých (poškozených) řádků pro W{TARGET_WEEK}")
kept = existing[~mask_to_remove]

combined = pd.concat([kept, new_df], ignore_index=True)
combined.to_csv(SNAPSHOTS_PATH, index=False)
print(f"Hotovo. Nový počet řádků celkem: {len(combined)}")
print(f"Týdny v souboru: {sorted(combined['week_label'].apply(extract_week_num).dropna().unique())}")
