from typing import Dict, List, Optional
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.utils import get_column_letter

NAVY="1F2A44"; ACCENT="2E6F9E"; ACCENT_LIGHT="DCE9F2"; GREEN="3D8B58"; RED="B23B3B"; LIGHT_GREY="F4F5F7"
GOLD="C9A227"; PLUM="6B4C7A"; GRID="E3E6EA"; MUTED="6B7280"; OLIVE="6E8B3D"; WEEKNUM_FILL="EDEFF2"
FN="Calibri"
TF=Font(name=FN,size=16,bold=True,color="FFFFFF"); STF=Font(name=FN,size=10,italic=True,color="FFFFFF")
HF=Font(name=FN,size=10,bold=True,color="FFFFFF"); LF=Font(name=FN,size=10,bold=True,color=NAVY)
CF=Font(name=FN,size=10,color="333333"); MF=Font(name=FN,size=10,bold=True,color=NAVY)
WF=Font(name=FN,size=9,italic=True,color=MUTED)
TFill=PatternFill("solid",fgColor=NAVY); HFill=PatternFill("solid",fgColor=ACCENT)
MFill=PatternFill("solid",fgColor=ACCENT_LIGHT); BFill=PatternFill("solid",fgColor=LIGHT_GREY)
WFill=PatternFill("solid",fgColor=WEEKNUM_FILL)
THIN=Side(style="thin",color="D0D3D8"); BDR=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
PCT={"Win rate (cumulative)"}
STAGE_PALETTE=[ACCENT,GREEN,GOLD,PLUM,RED,"5B8FB0","8FA998","D0A85C","9A7BAD","C77B7B","4C6B8A","7BA88F"]
MONEY_FMT='#,##0" Kč"'

# ---------------------------------------------------------------------------
# Chart appearance helpers - readable axes, sharp corners, consistent
# colors, consistent title/legend styling, no unnecessary borders.
# ---------------------------------------------------------------------------

def _hide_axis_labels(axis):
    """Makes axis tick labels invisible (white-on-white) instead of removing
    them outright - openpyxl's tickLblPos descriptor doesn't support the
    'none' value, so this is the reliable way to hide them."""
    axis.txPr = RichText(
        bodyPr=RichTextProperties(),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=100,solidFill="FFFFFF")),
                      endParaRPr=CharacterProperties(sz=100,solidFill="FFFFFF"))],
    )
    axis.majorTickMark = "none"

def _axis_label_style(axis, size=900, rotate=None):
    bodyPr = RichTextProperties(rot=int(rotate*60000)) if rotate is not None else RichTextProperties()
    axis.txPr = RichText(
        bodyPr=bodyPr,
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size,solidFill=MUTED)),
                      endParaRPr=CharacterProperties(sz=size,solidFill=MUTED))],
    )

def _thin_labels(axis, n_weeks, max_labels=16):
    """When there are many weeks, don't show every single label (they'd overlap)."""
    if n_weeks > max_labels:
        axis.tickLblSkip = max(1, round(n_weeks / max_labels))

def _sharp_lines(chart, width_pt=2.25):
    """Turns off smoothing on line charts - sharp, not rounded transitions,
    and thickens the line so it's clearly visible."""
    for s in chart.series:
        s.smooth = False
        s.graphicalProperties.line.width = int(width_pt*12700)

def _set_title(ch, text, size=1400, color=NAVY):
    ch.title = text
    try:
        run = ch.title.tx.rich.p[0].r[0]
        run.rPr = CharacterProperties(sz=size, b=True, solidFill=color, latin=None)
    except Exception:
        pass

def _no_border(ch):
    """Removes the border around the whole chart - cleaner, 'floating' look."""
    ch.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))

def _style_value_axis(ch, title=None, fmt=MONEY_FMT):
    if title: ch.y_axis.title = title
    ch.y_axis.numFmt = fmt
    if ch.y_axis.majorGridlines is not None:
        ch.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID, w=6350))
    _axis_label_style(ch.y_axis)
    ch.y_axis.delete = False

def _style_legend(ch, position="b", size=850):
    ch.legend = Legend(); ch.legend.position = position; ch.legend.overlay = False
    ch.legend.txPr = RichText(
        bodyPr=RichTextProperties(),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=size,solidFill=MUTED)), endParaRPr=None)],
    )

def _reserve_top_margin(ch, top=0.18, bottom_margin=0.06, side_margin=0.02):
    """
    Ručně vyhradí místo NAD plot area (pro titulek) a malé okraje po
    stranách/zdola - bez tohoto explicitního layoutu si to některé
    prohlížeče/verze Excelu spočítají špatně a popisky (zvlášť u funnelu,
    kde jsou dlouhé a mají vlastní data labels) se pak překrývají s
    titulkem grafu.

    Pozor: nastavuje se na ch.layout (ne ch.plot_area.layout) - openpyxl
    při ukládání přepíše plot_area.layout hodnotou z chart.layout, takže
    nastavení na plot_area by se tiše ztratilo.
    """
    ch.layout = Layout(manualLayout=ManualLayout(
        xMode="edge", yMode="edge",
        x=side_margin, y=top,
        w=1 - 2 * side_margin, h=1 - top - bottom_margin,
    ))

def _finish(ch, n_weeks, title, y_title="Kč", legend=True, legend_pos="b", rotate_x=-45, thin_x=True):
    """Shared final 'polish' step - called at the end of every chart."""
    _set_title(ch, title)
    _axis_label_style(ch.x_axis, rotate=rotate_x)
    if thin_x: _thin_labels(ch.x_axis, n_weeks)
    ch.x_axis.delete = False
    ch.gapWidth = 55
    _style_value_axis(ch, y_title)
    if legend: _style_legend(ch, legend_pos)
    else: ch.legend = None
    _no_border(ch)
    ch.roundedCorners = False
    ch.visible_cells_only = False
    _reserve_top_margin(ch)
    return ch

def _title(ws, title, subtitle, last_col):
    ws.merge_cells(f"A1:{last_col}1"); ws.merge_cells(f"A2:{last_col}2")
    ws["A1"]=title; ws["A1"].font=TF; ws["A1"].fill=TFill; ws["A1"].alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws["A2"]=subtitle; ws["A2"].font=STF; ws["A2"].fill=TFill; ws["A2"].alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=18

def write_table(ws, start_row, week_labels_display, week_nums, stage_order, stage_data, metrics_order, metrics):
    n = len(week_labels_display); r = start_row; header_row = r
    ws.cell(r,1,"Date").font=HF; ws.cell(r,1).fill=HFill; ws.cell(r,1).border=BDR
    for w,lbl in enumerate(week_labels_display):
        c=ws.cell(r,2+w,lbl); c.font=HF; c.fill=HFill; c.alignment=Alignment(horizontal="center"); c.border=BDR
    r += 1
    # Week number (ISO) row - right under the date, so it's clear which
    # week of the year this is (same numbering as the original excel).
    weeknum_row = r
    ws.cell(r,1,"Week #").font=WF; ws.cell(r,1).fill=WFill; ws.cell(r,1).border=BDR
    for w,wn in enumerate(week_nums):
        c=ws.cell(r,2+w,wn); c.font=WF; c.fill=WFill; c.alignment=Alignment(horizontal="center"); c.border=BDR
    r += 1
    stage_rows={}
    for i,stage in enumerate(stage_order):
        band=BFill if i%2==0 else PatternFill(fill_type=None)
        ws.cell(r,1,stage).font=CF; ws.cell(r,1).fill=band; ws.cell(r,1).border=BDR
        vals=stage_data.get(stage,[0]*n)
        for w in range(n):
            c=ws.cell(r,2+w,vals[w]); c.number_format="#,##0"; c.font=CF; c.fill=band; c.border=BDR; c.alignment=Alignment(horizontal="right")
        stage_rows[stage]=r; r+=1
    r+=1; metric_rows={}
    for name in metrics_order:
        vals=metrics.get(name)
        if vals is None: continue
        ws.cell(r,1,name).font=MF; ws.cell(r,1).fill=MFill; ws.cell(r,1).border=BDR
        for w in range(n):
            c=ws.cell(r,2+w,vals[w]); c.font=MF; c.fill=MFill; c.border=BDR; c.alignment=Alignment(horizontal="right")
            c.number_format="0.0%" if name in PCT else "#,##0"
        metric_rows[name]=r; r+=1
    for ci in range(1,2+n): ws.column_dimensions[get_column_letter(ci)].width=13 if ci>1 else 30
    return header_row, weeknum_row, stage_rows, metric_rows, r

def write_hidden_block(ws, start_row, n_cols, row_specs):
    """
    Generic helper - writes arbitrary numeric series into hidden rows
    (used as a chart data source without the user seeing it in Excel).
    row_specs = list of (label, values[n_cols]).
    Returns {label: row_index}, next_free_row.
    """
    r = start_row; rows = {}
    for label, values in row_specs:
        ws.cell(r,1,label)
        vals = values if values is not None else [None]*n_cols
        for w in range(n_cols):
            ws.cell(r,2+w, vals[w] if w < len(vals) else None)
        rows[label] = r; r += 1
    for rr in range(start_row, r):
        ws.row_dimensions[rr].hidden = True
    return rows, r

def write_raw_block(ws, start_row, n_weeks, stage_order, stage_data):
    """Hidden rows with stage x week values (raw or weighted - depending on
    what's passed in stage_data). Used as the data source for funnel
    charts. Plain stage names (no prefix), so chart legends/labels stay clean."""
    specs = [(stage, stage_data.get(stage,[0]*n_weeks)) for stage in stage_order]
    rows, next_r = write_hidden_block(ws, start_row, n_weeks, specs)
    return rows, next_r

def write_chart4_block(ws, start_row, chart4_data):
    """
    Hidden rows for chart 4 (Pace to goal) - the X axis spans the WHOLE
    YEAR (week 1..max_week), not just the displayed weeks. Actual metrics
    only have a value for weeks that already happened (elsewhere None ->
    the chart simply doesn't draw there, no fabricated data). Goal is a
    linear line for the whole year.
    """
    n = chart4_data["max_week"]
    specs = [
        ("Week", chart4_data["weeknums"]),
        ("Change in Rolling 18 (weekly)", chart4_data["changes_rolling18"]),
        ("Pipeline till end of year", chart4_data["pipeline_eoy"]),
        ("Rolling 18", chart4_data["rolling18"]),
        ("Won (cumulative)", chart4_data["won_cum"]),
        ("Goal (linear, full year)", chart4_data["goal_line"]),
    ]
    rows, next_r = write_hidden_block(ws, start_row, n, specs)
    return rows, next_r, n

def _add_charts(ws, anchor_row, header_row, metric_rows, weighted_eoy_rows, raw_full_rows, weighted_full_rows,
                 n_weeks, stage_order, week_dates_for_title=None, chart4_rows=None, chart4_n=None):
    mc = 1+n_weeks
    cats = Reference(ws,min_col=2,max_col=mc,min_row=header_row,max_row=header_row)
    last_week_lbl = week_dates_for_title[-1] if week_dates_for_title else ""

    def line(title, series_colors, y="Kč"):
        # series_colors: list of (metric_label, hex_color)
        ch=LineChart(); ch.height=10.5; ch.width=27
        for lbl,color in series_colors:
            if lbl not in metric_rows: continue
            d=Reference(ws,min_col=1,max_col=mc,min_row=metric_rows[lbl],max_row=metric_rows[lbl])
            ch.add_data(d,titles_from_data=True,from_rows=True)
        ch.set_categories(cats)
        for s,(lbl,color) in zip(ch.series, [x for x in series_colors if x[0] in metric_rows]):
            s.graphicalProperties.line.solidFill = color
            s.marker = Marker(symbol="circle", size=5)
            s.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color))
        _sharp_lines(ch)
        return _finish(ch, n_weeks, title, y)

    def bar(title, lbl, color=ACCENT, y="Kč"):
        ch=BarChart(); ch.type="col"; ch.height=10.5; ch.width=27
        d=Reference(ws,min_col=1,max_col=mc,min_row=metric_rows[lbl],max_row=metric_rows[lbl])
        ch.add_data(d,titles_from_data=True,from_rows=True); ch.set_categories(cats)
        ch.series[0].graphicalProperties.solidFill=color
        ch.series[0].graphicalProperties.line.noFill=True
        ch = _finish(ch, n_weeks, title, y, legend=False)
        ch.y_axis.title = None
        _hide_axis_labels(ch.y_axis)
        return ch

    def stacked_won_lost():
        # Cumulative (stacked) Won vs. Lost - better conveys the overall
        # split of closed pipeline over time than a weekly line.
        ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.height=10.5; ch.width=27
        for lbl,color in [("Won (cumulative)",GREEN),("Lost (cumulative)",RED)]:
            if lbl not in metric_rows: continue
            d=Reference(ws,min_col=1,max_col=mc,min_row=metric_rows[lbl],max_row=metric_rows[lbl])
            ch.add_data(d,titles_from_data=True,from_rows=True)
        ch.set_categories(cats)
        for s,color in zip(ch.series, [GREEN,RED]):
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
        ch = _finish(ch, n_weeks, "Won vs. Lost (cumulative)", "Kč", legend=True)
        ch.y_axis.title = None
        _hide_axis_labels(ch.y_axis)
        return ch

    def funnel_current():
        # Current pipeline breakdown (latest week) - ACTUAL (unweighted)
        # values, based on the Rolling 18 dataset (not "till end of year").
        # Left-to-right, first-stage-on-top, labelled directly on the bars
        # (no separate axis labels, no "[data]"/series-name clutter).
        ch=BarChart(); ch.type="bar"; ch.height=10.5; ch.width=27
        lc=1+n_weeks
        min_r, max_r = min(raw_full_rows.values()), max(raw_full_rows.values())
        d=Reference(ws,min_col=lc,max_col=lc,min_row=min_r,max_row=max_r)
        cs=Reference(ws,min_col=1,max_col=1,min_row=min_r,max_row=max_r)
        ch.add_data(d,titles_from_data=False); ch.set_categories(cs)
        ch.series[0].graphicalProperties.solidFill=GREEN
        ch.series[0].graphicalProperties.line.noFill=True
        ch.dataLabels=DataLabelList()
        ch.dataLabels.showVal=True; ch.dataLabels.showCatName=True
        ch.dataLabels.showSerName=False; ch.dataLabels.showLegendKey=False
        ch.dataLabels.showPercent=False; ch.dataLabels.showBubbleSize=False
        ch.dataLabels.separator=": "
        ch.dataLabels.numFmt='#,##0" Kč"'
        ch.dataLabels.txPr = RichText(bodyPr=RichTextProperties(),
            p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=850,b=True,solidFill=NAVY)), endParaRPr=None)])
        # First pipeline stage (Lead Engaged...) on top, not at the bottom.
        ch.x_axis.scaling.orientation = "maxMin"
        # Keep the value axis normal (0 on the left, ascending to the right)
        # and pinned to the BOTTOM - reversing the category axis above would
        # otherwise flip it up to the top by default.
        ch.y_axis.scaling.orientation = "minMax"
        ch.y_axis.crosses = "min"
        ch.y_axis.axPos = "b"
        sfx=f" ({last_week_lbl})" if last_week_lbl else ""
        ch = _finish(ch, n_weeks, f"Funnel – pipeline breakdown, Rolling 18{sfx}", "Kč", legend=False)
        ch.x_axis.title = None
        # Data labels on each bar already carry the stage name and amount -
        # hide the separate axis tick labels/titles on both axes to avoid
        # duplicate clutter.
        _hide_axis_labels(ch.x_axis)
        ch.y_axis.title = None
        _hide_axis_labels(ch.y_axis)
        # Data labels here can wrap to two lines (long stage names) - give
        # a bit more breathing room below the title than the shared default.
        _reserve_top_margin(ch, top=0.22)
        return ch

    def funnel_evolution(title, rows_dict, color_offset=0):
        # Stacked column chart: evolution of the WEIGHTED funnel over time,
        # week by week, from the start of the year - one series per stage.
        ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.height=11; ch.width=27
        d=Reference(ws,min_col=1,max_col=mc,min_row=min(rows_dict.values()),max_row=max(rows_dict.values()))
        ch.add_data(d,titles_from_data=True,from_rows=True); ch.set_categories(cats)
        for i,s in enumerate(ch.series):
            color = STAGE_PALETTE[(i+color_offset)%len(STAGE_PALETTE)]
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
        return _finish(ch, n_weeks, title, "Kč", legend=True)

    def combo_tempo():
        # "Pace to goal" - the X axis spans the WHOLE YEAR (week 1..max_week),
        # not just the displayed weeks. Columns = weekly change in Rolling 18
        # (secondary axis), lines = Pipeline till end of year, Rolling 18,
        # Won cumulative (primary axis, only where data actually exists), and
        # a linear Goal line for the whole year (annual_goal / weeks in year).
        if not chart4_rows: return None
        n4 = chart4_n; mc4 = 1+n4
        cats4 = Reference(ws,min_col=2,max_col=mc4,min_row=chart4_rows["Week"],max_row=chart4_rows["Week"])

        bar_ch=BarChart(); bar_ch.type="col"
        d=Reference(ws,min_col=1,max_col=mc4,min_row=chart4_rows["Change in Rolling 18 (weekly)"],max_row=chart4_rows["Change in Rolling 18 (weekly)"])
        bar_ch.add_data(d,titles_from_data=True,from_rows=True); bar_ch.set_categories(cats4)
        bar_ch.series[0].graphicalProperties.solidFill=ACCENT_LIGHT
        bar_ch.series[0].graphicalProperties.line.solidFill=ACCENT
        bar_ch.y_axis.axId=200; bar_ch.y_axis.title="Weekly change (Kč)"; bar_ch.y_axis.crosses="max"
        bar_ch.y_axis.numFmt=MONEY_FMT; _axis_label_style(bar_ch.y_axis)
        bar_ch.y_axis.majorGridlines=None

        line_ch=LineChart()
        series_specs=[("Pipeline till end of year",GOLD),("Rolling 18",ACCENT),("Won (cumulative)",GREEN)]
        for name,color in series_specs:
            dd=Reference(ws,min_col=1,max_col=mc4,min_row=chart4_rows[name],max_row=chart4_rows[name])
            line_ch.add_data(dd,titles_from_data=True,from_rows=True)
        for s,(name,color) in zip(line_ch.series, series_specs):
            s.graphicalProperties.line.solidFill = color
            s.marker = Marker(symbol="circle", size=5)
            s.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color))
        # Goal - a standalone linear line for the whole year, no markers,
        # dashed, so it's visually clear this is a TARGET, not a measurement.
        goal_d=Reference(ws,min_col=1,max_col=mc4,min_row=chart4_rows["Goal (linear, full year)"],max_row=chart4_rows["Goal (linear, full year)"])
        line_ch.add_data(goal_d,titles_from_data=True,from_rows=True)
        goal_series = line_ch.series[-1]
        goal_series.graphicalProperties.line.solidFill = OLIVE
        goal_series.graphicalProperties.line.dashStyle = "dash"
        goal_series.marker = Marker(symbol="none")
        line_ch.set_categories(cats4)
        _sharp_lines(line_ch)
        line_ch.y_axis.axId=100; line_ch.y_axis.title="Cumulative value (Kč)"
        line_ch.y_axis.numFmt=MONEY_FMT; _axis_label_style(line_ch.y_axis)
        if line_ch.y_axis.majorGridlines is not None:
            line_ch.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID, w=6350))

        bar_ch += line_ch
        # Taller than the others, so smaller differences (e.g. in Won) are
        # easier to see across the full-year Y axis range.
        bar_ch.height=15; bar_ch.width=27
        ch = _finish(bar_ch, n4, "Pace to goal – pipeline, Rolling 18, Won and Goal for the full year",
                     y_title=None, legend=True, thin_x=True)
        ch.y_axis.numFmt=MONEY_FMT
        return ch

    chart1 = bar("Weekly change in Rolling 18","Changes in Rolling 18", color=ACCENT)
    chart2 = stacked_won_lost()
    chart3 = funnel_current()
    chart4 = combo_tempo()
    chart5 = funnel_evolution("Funnel evolution over time – till end of year (weighted values)", weighted_eoy_rows)
    chart6 = funnel_evolution("Funnel evolution over time – Rolling 18 (weighted values)", weighted_full_rows)

    # Layout: chart4 sits to the right of chart1, chart5 to the right of
    # chart2, chart6 to the right of chart3 - two side-by-side columns
    # instead of one long stack. LEFT_COL is nudged right of column A (away
    # from the narrow label column edge). RIGHT_COL is far enough right
    # (~49cm) that it never overlaps the ~27cm-wide left-column charts even
    # after that nudge.
    LEFT_COL = "C"
    RIGHT_COL = "S"
    r = anchor_row
    ws.add_chart(chart1, f"{LEFT_COL}{r}")
    if chart4 is not None:
        ws.add_chart(chart4, f"{RIGHT_COL}{r}")
    r += 30  # chart4 (15cm tall) needs more vertical room than chart1
    ws.add_chart(chart2, f"{LEFT_COL}{r}")
    ws.add_chart(chart5, f"{RIGHT_COL}{r}")
    r += 23
    ws.add_chart(chart3, f"{LEFT_COL}{r}")
    ws.add_chart(chart6, f"{RIGHT_COL}{r}")
    r += 23
    return r

METRICS_ORDER=["Won","Lost","Pipeline till end of year","Changes in pipeline till end of the year",
               "Rolling 18","Changes in Rolling 18",
               "Win rate (cumulative)","Avg. deal size","Won (cumulative)","Lost (cumulative)","Goal (cumulative)"]

def build_person_sheet(wb, owner, sheet_data, stage_order, week_labels_display, week_dates_for_title, week_nums):
    ws=wb.create_sheet(owner[:31]); n=len(week_labels_display); lc=get_column_letter(1+n)
    goal=sheet_data.get("annual_goal")
    goal_txt=f"Annual goal: {goal:,.0f} Kč".replace(",","_").replace("_"," ") if goal else "Annual goal: not yet set"
    _title(ws,f"{owner} — Sales Pipeline Report",f"Weekly pipeline overview, Kč  ·  {goal_txt}",lc)
    hr,wnr,sr,mr,er=write_table(ws,4,week_labels_display,week_nums,stage_order,sheet_data["stage_weighted"],METRICS_ORDER,sheet_data["metrics"])
    raw_full_rows, er2 = write_raw_block(ws, er, n, stage_order, sheet_data.get("raw_full", {}))
    weighted_full_rows, er3 = write_raw_block(ws, er2, n, stage_order, sheet_data.get("weighted_full", {}))
    chart4_rows, er4, chart4_n = (None, er3, None)
    if sheet_data.get("chart4"):
        chart4_rows, er4, chart4_n = write_chart4_block(ws, er3, sheet_data["chart4"])
    _add_charts(ws,er4+2,hr,mr,sr,raw_full_rows,weighted_full_rows,n,stage_order,week_dates_for_title,chart4_rows,chart4_n)
    ws.freeze_panes="B5"; ws.sheet_view.showGridLines=False

def build_aggregation_sheet(wb, agg_data, stage_order, week_labels_display, week_dates_for_title, leaderboard, week_nums):
    ws=wb.create_sheet("Aggregation"); n=len(week_labels_display); lc=get_column_letter(1+n)
    _title(ws,"Tapix — Aggregated Sales Pipeline Report","All sales reps combined, Kč",lc)
    hr,wnr,sr,mr,er=write_table(ws,4,week_labels_display,week_nums,stage_order,agg_data["stage_weighted"],METRICS_ORDER,agg_data["metrics"])
    raw_full_rows, er2 = write_raw_block(ws, er, n, stage_order, agg_data.get("raw_full", {}))
    weighted_full_rows, er3 = write_raw_block(ws, er2, n, stage_order, agg_data.get("weighted_full", {}))
    chart4_rows, er4, chart4_n = (None, er3, None)
    if agg_data.get("chart4"):
        chart4_rows, er4, chart4_n = write_chart4_block(ws, er3, agg_data["chart4"])
    end=_add_charts(ws,er4+2,hr,mr,sr,raw_full_rows,weighted_full_rows,n,stage_order,week_dates_for_title,chart4_rows,chart4_n)
    lb=end+2; ws.cell(lb,1,"Sales rep leaderboard (Total Won)").font=LF; lb+=1
    ws.cell(lb,1,"Sales rep").font=HF; ws.cell(lb,1).fill=HFill
    ws.cell(lb,2,"Total Won (Kč)").font=HF; ws.cell(lb,2).fill=HFill
    lbh=lb; lb+=1; lbf=lb
    for name,total in leaderboard:
        ws.cell(lb,1,name).font=CF
        c=ws.cell(lb,2,round(total)); c.number_format="#,##0"; c.font=CF; lb+=1
    lbl=max(lb-1,lbf)
    ch=BarChart(); ch.type="bar"; ch.height=10.5; ch.width=27
    d=Reference(ws,min_col=2,max_col=2,min_row=lbh,max_row=lbl)
    cs=Reference(ws,min_col=1,max_col=1,min_row=lbf,max_row=lbl)
    ch.add_data(d,titles_from_data=True); ch.set_categories(cs)
    ch.series[0].graphicalProperties.solidFill=ACCENT
    ch.series[0].graphicalProperties.line.noFill=True
    ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True; ch.dataLabels.numFmt=MONEY_FMT
    ch.y_axis.scaling.orientation="maxMin"
    ch = _finish(ch, len(leaderboard) or 1, "Sales rep leaderboard by total Won", "Kč", legend=False)
    ws.add_chart(ch,f"D{lbh}"); ws.freeze_panes="B5"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=30

def build_ledger_sheet(wb, title, rows, color):
    ws=wb.create_sheet(title)
    _title(ws,f"{title} Deals — automatic ledger","Generated weekly from HubSpot","G")
    headers=["Deal ID","Deal Name","Company","Deal Owner","Close Date","Amount (Kč)","Week #"]
    r=4
    for ci,h in enumerate(headers):
        c=ws.cell(r,1+ci,h); c.font=HF; c.fill=PatternFill("solid",fgColor=color); c.border=BDR
    r+=1
    for row in sorted(rows, key=lambda x: x[4] if x[4] else date.min):
        for ci,val in enumerate(row):
            c=ws.cell(r,1+ci,val); c.font=CF; c.border=BDR
            if ci==4: c.number_format="dd.mm.yyyy"
            if ci==5: c.number_format="#,##0"
            if ci==6: c.alignment=Alignment(horizontal="center")
        r+=1
    for ci,w in enumerate([10,26,16,14,13,14,10]): ws.column_dimensions[get_column_letter(1+ci)].width=w
    ws.freeze_panes="A5"; ws.sheet_view.showGridLines=False

def build_raw_debug_sheet(wb, title, raw_by_owner, stage_order, week_labels_display, owners):
    ws = wb.create_sheet(title[:31])
    n = len(week_labels_display)
    last_col = get_column_letter(1 + n)
    _title(ws, f"DEBUG: {title}", "Temporary - raw unweighted data, for comparison with the old system", last_col)

    r = 4
    for owner in owners:
        c = ws.cell(r, 1, f"— {owner} —")
        c.font = LF
        c.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
        r += 1
        header_row = r
        ws.cell(r, 1, "Stage").font = HF
        ws.cell(r, 1).fill = HFill
        for w, lbl in enumerate(week_labels_display):
            cc = ws.cell(r, 2 + w, lbl)
            cc.font = HF; cc.fill = HFill; cc.alignment = Alignment(horizontal="center")
        r += 1
        owner_data = raw_by_owner.get(owner, {})
        for stage in stage_order:
            ws.cell(r, 1, stage).font = CF
            vals = owner_data.get(stage, [0] * n)
            for w in range(n):
                cc = ws.cell(r, 2 + w, vals[w])
                cc.number_format = "#,##0"
                cc.font = CF
                cc.alignment = Alignment(horizontal="right")
            r += 1
        r += 1

    for ci in range(1, 2 + n):
        ws.column_dimensions[get_column_letter(ci)].width = 13 if ci > 1 else 32
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    return ws


def new_workbook():
    wb=Workbook(); wb.remove(wb.active); return wb
