#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Jednorázová oprava historických dat W9-W27 na základě syrových exportů
z obou starých skriptů (HubSpot_Deals_By_Stage_2026.xlsx = fixed cutoff,
HubSpot_Deals_By_Stage_DYNAMIC_2026.xlsx = 18měsíční cutoff).

- Nové fáze (Not a lead, To be contacted, Qualified -> Deal) se VYNECHÁVAJÍ.
- Won/Lost se taky nepřenáší (ty se řeší živě, ne z historie).
- W1-W8 se nedotýkají. W28+ (live data) se nedotýkají.
"""
import re
import sys
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, "/home/claude/repo/src")
import data_store as store

FIXED_PATH = "/mnt/user-data/uploads/HubSpot_Deals_By_Stage_2026__3_.xlsx"
DYNAMIC_PATH = "/mnt/user-data/uploads/HubSpot_Deals_By_Stage_DYNAMIC_2026.xlsx"
SNAPSHOTS_PATH = Path("/home/claude/repo/data/deals_snapshots.csv")

CANONICAL_STAGES = [
    "Lead Engaged", "Meeting negotiation / contacted", "Intro meeting agreed",
    "Awaiting confirmation / Not Now", "Qualify", "Discover", "Validate",
    "Decide", "Commit", "Tech intro", "Implementation", "Testing",
]

SHEET_TO_TARGET_OWNER = {
    "Martin Korbelar": "Martin Korbelar",
    "Lukas Hora": "Lukas Hora",
    "Veronika Kincová": "Veronika Kincová",
    "Mykyta Artiukhov": "Mykyta Artiukhov",
    "Richard Brůža": "Richard Brůža",
    "Oldřich Huzil": "Oldřich Huzil",
    "Julie Mrkvičková": "Julie Mrkvickova",   # stará (do W15), sjednotit na aktuální jméno
    "Julie Mrkvickova": "Julie Mrkvickova",   # nová (od W16)
    "Dominika Šudová": "Dominika Šudová",
}

WEEK_MIN, WEEK_MAX = 9, 27


def extract_week_num(header):
    m = re.match(r"2026-W(\d+)", header or "")
    return int(m.group(1)) if m else None


def read_raw_file(path):
    """Vrací {(owner, stage, week_label): amount}"""
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        target_owner = SHEET_TO_TARGET_OWNER.get(sheet)
        if not target_owner:
            continue
        ws = wb[sheet]
        stage_rows = {}
        for r in range(2, ws.max_row + 1):
            lbl = ws.cell(r, 1).value
            if lbl:
                stage_rows[lbl.strip()] = r
        for c in range(2, ws.max_column + 1):
            header = ws.cell(1, c).value
            wn = extract_week_num(header)
            if wn is None or not (WEEK_MIN <= wn <= WEEK_MAX):
                continue
            for stage in CANONICAL_STAGES:
                r = stage_rows.get(stage)
                if r is None:
                    continue
                val = ws.cell(r, c).value
                amt = float(val) if isinstance(val, (int, float)) else 0.0
                out[(target_owner, stage, header)] = out.get((target_owner, stage, header), 0.0) + amt
    return out


print("Načítám FIXED (till end of year)...")
fixed = read_raw_file(FIXED_PATH)
print(f"  {len(fixed)} (owner, stage, week) záznamů")

print("Načítám DYNAMIC (Rolling 18)...")
dynamic = read_raw_file(DYNAMIC_PATH)
print(f"  {len(dynamic)} (owner, stage, week) záznamů")

all_keys = set(fixed.keys()) | set(dynamic.keys())
print(f"Celkem unikátních (owner, stage, week): {len(all_keys)}")

new_rows = []
negative_deltas = 0
for owner, stage, week_label in all_keys:
    fixed_amt = fixed.get((owner, stage, week_label), 0.0)
    dynamic_amt = dynamic.get((owner, stage, week_label), 0.0)
    delta = dynamic_amt - fixed_amt
    if delta < -0.01:
        negative_deltas += 1
        delta = 0.0

    wn = extract_week_num(week_label)
    monday = date.fromisocalendar(2026, wn, 1)

    if fixed_amt != 0 or delta == 0:
        new_rows.append({
            "week_label": week_label, "week_monday": monday.isoformat(),
            "deal_id": f"hist2-{owner}-{stage}-{week_label}-A".replace(" ", "_"),
            "deal_name": "", "owner_id": "", "owner_name": owner,
            "stage_id": stage, "stage_label": stage, "amount": fixed_amt,
            "closedate": monday.isoformat(), "pipeline_id": "",
            "is_closed_won": False, "is_closed_lost": False,
        })
    if delta > 0:
        # Mimo "do konce roku" okno, ale v 18měsíčním okně -> jen do Rolling 18
        beyond_eoy_date = date(2027, 6, 1)
        new_rows.append({
            "week_label": week_label, "week_monday": monday.isoformat(),
            "deal_id": f"hist2-{owner}-{stage}-{week_label}-B".replace(" ", "_"),
            "deal_name": "", "owner_id": "", "owner_name": owner,
            "stage_id": stage, "stage_label": stage, "amount": delta,
            "closedate": beyond_eoy_date.isoformat(), "pipeline_id": "",
            "is_closed_won": False, "is_closed_lost": False,
        })

print(f"Záporné delty (dynamic < fixed, ořezáno na 0): {negative_deltas}")
print(f"Nově sestavených řádků: {len(new_rows)}")

new_df = pd.DataFrame(new_rows, columns=store.SNAPSHOTS_COLUMNS)

# Načti stávající data a odeber W9-W27, W1-W8 a W28+ nech beze změny
existing = pd.read_csv(SNAPSHOTS_PATH)
existing_week_nums = existing["week_label"].apply(extract_week_num)
mask_to_remove = existing_week_nums.between(WEEK_MIN, WEEK_MAX)
print(f"Odebírám {mask_to_remove.sum()} starých řádků pro W{WEEK_MIN}-W{WEEK_MAX}")
kept = existing[~mask_to_remove]

combined = pd.concat([kept, new_df], ignore_index=True)
combined.to_csv(SNAPSHOTS_PATH, index=False)
print(f"Hotovo. Nový počet řádků celkem: {len(combined)}")
print(f"Týdny v souboru: {sorted(combined['week_label'].apply(extract_week_num).dropna().unique())}")
